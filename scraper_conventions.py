"""
Scraper - Conventions de stage OMNES Education (Selenium + login ADFS)
=======================================================================
Dépendances : pip install selenium openpyxl
"""

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import time
import re

# ─── CONFIGURATION ────────────────────────────────────────────────────────────
EMAIL       = "ilias.messaoudi@edu.ece.fr"       # ← Remplace par ton email OMNES
PASSWORD    = "XJ8cArbc_"          # ← Remplace par ton mot de passe

BASE_URL    = "https://convention.omneseducation.com"
LIST_URL    = f"{BASE_URL}/historique-conventions.html"
OUTPUT_FILE = "conventions_emails.xlsx"
PAGE_DELAY  = 3.0
# ──────────────────────────────────────────────────────────────────────────────


def init_driver() -> webdriver.Chrome:
    options = Options()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument("--window-size=1400,900")
    driver = webdriver.Chrome(options=options)
    return driver


def login(driver: webdriver.Chrome) -> bool:
    """Navigue vers le site, remplit le formulaire ADFS automatiquement."""
    print("🌐 Navigation vers le site...")
    driver.get(LIST_URL)
    time.sleep(3)

    # Si on est redirigé vers ADFS (page de login)
    if "adfs" in driver.current_url or "login" in driver.current_url.lower():
        print(f"🔐 Page ADFS détectée → {driver.current_url[:80]}...")
        try:
            # Champ email
            email_field = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='email'], input[name='UserName'], input[id='userNameInput'], input[type='text']"))
            )
            email_field.clear()
            email_field.send_keys(EMAIL)
            print(f"  ✅ Email saisi : {EMAIL}")
            time.sleep(0.5)

            # Champ mot de passe
            pwd_field = driver.find_element(By.CSS_SELECTOR, "input[type='password'], input[name='Password'], input[id='passwordInput']")
            pwd_field.clear()
            pwd_field.send_keys(PASSWORD)
            print("  ✅ Mot de passe saisi")
            time.sleep(0.5)

            # Bouton connexion
            submit = driver.find_element(By.CSS_SELECTOR, "input[type='submit'], button[type='submit'], #submitButton")
            submit.click()
            print("  ✅ Connexion envoyée")

            # Attente de la redirection vers le site
            WebDriverWait(driver, 20).until(
                lambda d: "convention.omneseducation.com" in d.current_url
            )
            time.sleep(PAGE_DELAY)
            print(f"  ✅ Connecté ! URL : {driver.current_url}")
            return True

        except TimeoutException:
            print("  ❌ Timeout pendant le login")
            return False
        except NoSuchElementException as e:
            print(f"  ❌ Champ non trouvé : {e}")
            return False

    elif "convention.omneseducation.com" in driver.current_url:
        print(f"✅ Déjà connecté ! URL : {driver.current_url}")
        return True

    print(f"⚠️ URL inattendue : {driver.current_url}")
    return False


def wait_for_conventions(driver: webdriver.Chrome, timeout: int = 15) -> bool:
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.listing-container"))
        )
        return True
    except TimeoutException:
        return False


def get_total_pages(driver: webdriver.Chrome) -> int:
    pages = []
    try:
        for a in driver.find_elements(By.CSS_SELECTOR, "a[href*='page=']"):
            href = a.get_attribute("href") or ""
            m = re.search(r"page=(\d+)", href)
            if m:
                pages.append(int(m.group(1)))
    except Exception:
        pass
    return max(pages) if pages else 1


def parse_current_page(driver: webdriver.Chrome, page_num: int) -> list[dict]:
    results = []

    if not wait_for_conventions(driver):
        print(f"  ❌ Pas de listing-container sur page {page_num}")
        with open(f"debug_page_{page_num}.html", "w", encoding="utf-8") as f:
            f.write(driver.page_source)
        print(f"  📄 HTML → debug_page_{page_num}.html")
        return []

    blocs = driver.find_elements(By.CSS_SELECTOR, "div.listing-container")
    print(f"  📦 {len(blocs)} conventions sur page {page_num}")

    for bloc in blocs:
        try:
            # Entreprise
            entreprise = ""
            try:
                titre = bloc.find_element(By.CSS_SELECTOR, "div.listing-row-title, h2.listing-row-title")
                entreprise = re.sub(r"\s*\(.*?\)\s*$", "", titre.text.strip()).strip()
            except NoSuchElementException:
                pass

            # Ville
            ville = ""
            try:
                small = bloc.find_element(By.CSS_SELECTOR, "div.listing-row-title small")
                ville = small.text.strip()
            except NoSuchElementException:
                m = re.search(r"\d{5}\s+([A-ZÉÈÀÙÂÊÎÔÛÇ][A-Z\s\-]{2,30})", bloc.text)
                if m:
                    ville = m.group(0).strip()

            # Adresse
            adresse = ""
            try:
                for div in bloc.find_elements(By.CSS_SELECTOR, "div.listing-row-content"):
                    txt = div.text.strip()
                    if "Siège" in txt or re.search(r"\d{1,3}\s+\w+.*\d{5}", txt):
                        adresse = txt
                        break
            except Exception:
                pass

            # Tuteur + Email
            nom = ""
            email = ""
            try:
                for small in bloc.find_elements(By.CSS_SELECTOR, "div.listing-row-content small"):
                    txt = small.text.strip()
                    if "Tuteur" in txt or "@" in txt:
                        emails = re.findall(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", txt)
                        if emails:
                            email = emails[0]
                        pattern = r"Tuteur[^:]*:\s*(.+?)\s*-\s*" + re.escape(email) if email else r"Tuteur[^:]*:\s*(.{5,50})"
                        m = re.search(pattern, txt)
                        if m:
                            nom = m.group(1).strip()
                        break
            except Exception:
                pass

            # Étudiant
            etudiant = ""
            try:
                for div in bloc.find_elements(By.CSS_SELECTOR, "div.listing-row-content"):
                    bold = div.find_elements(By.CSS_SELECTOR, "strong, b")
                    if bold:
                        etudiant = bold[0].text.strip()
                        break
            except Exception:
                pass

            # Type de stage + Période
            type_stage = ""
            periode = ""
            bloc_text = bloc.text
            for kw in ["Stage de fin d'études", "Alternance", "Stage classique", "Stage"]:
                if kw.lower() in bloc_text.lower():
                    type_stage = kw
                    break
            dates = re.findall(r"\d{2}/\d{2}/\d{4}", bloc_text)
            if len(dates) >= 2:
                periode = f"{dates[0]} > {dates[1]}"

            # Mission
            mission = ""
            try:
                mission_div = bloc.find_element(By.CSS_SELECTOR, "div[id^='stage-mission-']")
                mission = mission_div.text.strip()[:500]
            except NoSuchElementException:
                pass

            # Lien détail
            lien = ""
            try:
                for a in bloc.find_elements(By.TAG_NAME, "a"):
                    href = a.get_attribute("href") or ""
                    txt  = a.text.lower()
                    if "détail" in txt or "detail" in txt or re.search(r"/convention/\d|/stage/\d", href):
                        lien = href
                        break
            except Exception:
                pass

            results.append({
                "Étudiant":      etudiant,
                "Nom tuteur":    nom,
                "Email":         email,
                "Entreprise":    entreprise,
                "Ville":         ville,
                "Adresse":       adresse,
                "Type de stage": type_stage,
                "Mission":       mission,
                "Période":       periode,
                "Lien détail":   lien,
                "Source":        f"Page {page_num}",
            })

        except Exception as e:
            print(f"  ⚠️ Erreur bloc : {e}")
            continue

    return results


def enrich_detail(driver: webdriver.Chrome, entry: dict) -> None:
    url = entry.get("Lien détail", "")
    if not url:
        return
    try:
        driver.get(url)
        time.sleep(2)

        for a in driver.find_elements(By.TAG_NAME, "a"):
            href = a.get_attribute("href") or ""
            if "linkedin.com/in" in href and not entry.get("LinkedIn"):
                entry["LinkedIn"] = href
            elif ("twitter.com/" in href or "x.com/" in href) and not entry.get("Twitter"):
                entry["Twitter"] = href

        if not entry.get("Mission"):
            try:
                m = driver.find_element(By.CSS_SELECTOR, "div[id^='stage-mission-']")
                entry["Mission"] = m.text.strip()[:500]
            except NoSuchElementException:
                pass

        driver.back()
        time.sleep(1.5)

    except Exception as e:
        print(f"    ⚠️ {e}")
        driver.get(LIST_URL)
        time.sleep(2)


def build_excel(data: list[dict], filename: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Conventions"

    colonnes = [
        "Étudiant", "Nom tuteur", "Email", "Entreprise", "Ville", "Adresse",
        "Type de stage", "Mission", "LinkedIn", "Twitter",
        "Période", "Lien détail", "Source",
    ]

    for col_idx, col_name in enumerate(colonnes, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill      = PatternFill("solid", start_color="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    alt_fill = PatternFill("solid", start_color="EBF0FA")
    for row_idx, entry in enumerate(data, 2):
        for col_idx, col_name in enumerate(colonnes, 1):
            val  = entry.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font      = Font(name="Arial", size=10)
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            if col_name in ("Lien détail", "LinkedIn", "Twitter") and val:
                cell.font      = Font(name="Arial", size=10, color="1155CC", underline="single")
                cell.hyperlink = val

    widths = {
        "Étudiant": 22, "Nom tuteur": 25, "Email": 36, "Entreprise": 28,
        "Ville": 20, "Adresse": 35, "Type de stage": 22, "Mission": 50,
        "LinkedIn": 35, "Twitter": 30, "Période": 22, "Lien détail": 35, "Source": 12,
    }
    for col_idx, col_name in enumerate(colonnes, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 20)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(colonnes))}{len(data) + 1}"
    ws.freeze_panes    = "A2"

    ws2 = wb.create_sheet("Stats")
    for i, (k, v) in enumerate([
        ("Date extraction",   datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Total conventions", len(data)),
        ("Avec email",        sum(1 for d in data if d.get("Email"))),
        ("Avec LinkedIn",     sum(1 for d in data if d.get("LinkedIn"))),
    ], 1):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=True, name="Arial")
        ws2.cell(row=i, column=2, value=v).font = Font(name="Arial")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 30

    wb.save(filename)
    print(f"\n✅ Fichier généré : {filename}")
    print(f"   {len(data)} conventions | {sum(1 for d in data if d.get('Email'))} emails")


def main():
    print("🚀 Lancement de Chrome...")
    driver = init_driver()

    try:
        if not login(driver):
            print("❌ Échec de connexion.")
            input("Appuie sur Entrée pour fermer...")
            return

        driver.get(LIST_URL)
        time.sleep(PAGE_DELAY)

        total = get_total_pages(driver)
        print(f"\n📄 {total} page(s) détectée(s)\n")

        all_data = parse_current_page(driver, 1)

        for page in range(2, total + 1):
            print(f"\n📄 Page {page}/{total}...")
            driver.get(f"{LIST_URL}?page={page}")
            time.sleep(PAGE_DELAY)
            all_data.extend(parse_current_page(driver, page))

        print(f"\n📊 Total : {len(all_data)} conventions")
        print("🔗 Enrichissement...\n")

        for i, entry in enumerate(all_data, 1):
            if entry.get("Lien détail"):
                print(f"  [{i}/{len(all_data)}] {entry.get('Entreprise', '?')[:40]}")
                enrich_detail(driver, entry)

        build_excel(all_data, OUTPUT_FILE)

    finally:
        driver.quit()
        print("🔒 Chrome fermé.")


if __name__ == "__main__":
    main()